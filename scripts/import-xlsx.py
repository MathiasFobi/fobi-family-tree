#!/usr/bin/env python3
"""
Fobi Family Tree — spreadsheet → family.json converter.

Usage:
    python3 scripts/import-xlsx.py [path/to/spreadsheet.xlsx]

Default path: ./Vincent_and_Barbara_Fobi_Descendants.xlsx (in the project root)

Reads the 5-column spreadsheet:
  Children | Grand Children | Great Grand Children | (middle) | (surname)

Generates: data/family.json

Re-run this whenever the spreadsheet is updated, then commit + push.
"""

import json
import os
import re
import sys
from collections import OrderedDict

import openpyxl

DEFAULT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Vincent_and_Barbara_Fobi_Descendants.xlsx",
)


def slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")
    return s


def clean_name(name):
    if not name:
        return ""
    return re.sub(r"\?+", "", str(name)).strip()


def is_rip(name):
    return bool(name) and "rip" in str(name).lower()


def status_of(name):
    return "remembered" if is_rip(name) else "living"


def build_family(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]

    # Skip rows 1-2 (title and column headers)
    gen1_map: OrderedDict[str, dict] = OrderedDict()
    gen2_map: OrderedDict[tuple, dict] = OrderedDict()
    gen3_list: list[dict] = []

    current_gen1 = None
    current_gen2 = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        c1, c2, c3, c4, c5 = [str(v).strip() if v is not None else "" for v in row]

        # Gen 1 (column 0)
        if c1:
            current_gen1 = c1
            if c1 not in gen1_map:
                gen1_map[c1] = {
                    "id": slugify(c1),
                    "name": clean_name(c1),
                    "maidenName": None,
                    "spouse": None,
                    "status": status_of(c1),
                    "generation": 1,
                    "children": [],
                }
            current_gen2 = None

        # Gen 2 (column 1)
        if c2:
            current_gen2 = c2
            key = (current_gen1, c2)
            if key not in gen2_map:
                entry = {
                    "id": slugify(c2),
                    "name": clean_name(c2),
                    "middleName": None,
                    "surname": None,
                    "status": status_of(c2),
                    "generation": 2,
                    "parentId": gen1_map[current_gen1]["id"] if current_gen1 else None,
                    "children": [],
                }
                gen2_map[key] = entry
                if current_gen1:
                    gen1_map[current_gen1]["children"].append(entry["id"])

        # Gen 3 (column 2)
        if c3:
            eid = slugify(c3)
            gen3_list.append(
                {
                    "id": eid,
                    "name": clean_name(c3),
                    "middleName": clean_name(c4) if c4 else None,
                    "surname": clean_name(c5) if c5 else None,
                    "status": status_of(c3),
                    "generation": 3,
                    "parentId": (
                        gen2_map[(current_gen1, current_gen2)]["id"]
                        if (current_gen1, current_gen2) in gen2_map
                        else None
                    ),
                }
            )
            if (current_gen1, current_gen2) in gen2_map:
                gen2_map[(current_gen1, current_gen2)]["children"].append(eid)

    # Second pass: fill middle/surname for Gen 2 from the row where they appear
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        c1, c2, c3, c4, c5 = [str(v).strip() if v is not None else "" for v in row]
        if c2 and (c1, c2) in gen2_map:
            if c4 and not gen2_map[(c1, c2)]["middleName"]:
                gen2_map[(c1, c2)]["middleName"] = clean_name(c4)
            if c5 and not gen2_map[(c1, c2)]["surname"]:
                gen2_map[(c1, c2)]["surname"] = clean_name(c5)

    remembered = (
        sum(1 for e in gen1_map.values() if e["status"] == "remembered")
        + sum(1 for e in gen2_map.values() if e["status"] == "remembered")
        + sum(1 for e in gen3_list if e["status"] == "remembered")
    )

    return {
        "ancestors": [
            {"name": "Vincent Fobi", "role": "Patriarch"},
            {"name": "Barbara Fobi", "role": "Matriarch"},
        ],
        "stats": {
            "gen1": len(gen1_map),
            "gen2": len(gen2_map),
            "gen3": len(gen3_list),
            "remembered": remembered,
            "total": len(gen1_map) + len(gen2_map) + len(gen3_list),
        },
        "generations": {
            "1": list(gen1_map.values()),
            "2": list(gen2_map.values()),
            "3": gen3_list,
        },
    }


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    if not os.path.exists(path):
        print(f"✗ File not found: {path}", file=sys.stderr)
        print(
            f"  Usage: python3 {sys.argv[0]} [path/to/spreadsheet.xlsx]",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"→ Reading {path}...")
    family = build_family(path)

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "data",
        "family.json",
    )
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(family, f, indent=2, ensure_ascii=False)

    s = family["stats"]
    print(f"✓ Wrote {out_path}")
    print(f"  Gen 1:        {s['gen1']}")
    print(f"  Gen 2:        {s['gen2']}")
    print(f"  Gen 3:        {s['gen3']}")
    print(f"  Remembered:   {s['remembered']}")
    print(f"  Total:        {s['total']}")


if __name__ == "__main__":
    main()
